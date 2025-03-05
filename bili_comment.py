import requests
import re
import time
import os
from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

def get_popular_videos():
    # API地址
    url = "https://api.bilibili.com/x/web-interface/popular"
    read_list=[]

    try:
        # 发送GET请求
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # 检查请求是否成功
        
        data = response.json()
        
        # 提取视频列表
        video_list = data.get('data', {}).get('list', [])

        for video in video_list:
            title = video.get('title', '')
            short_link = video.get('short_link_v2', '')
            
            # 从短链接中提取BV号
            bv_id = urlparse(short_link).path.split('/')[-1] if short_link else ''

            read_list.append((title,bv_id))
        
        return read_list
        
    except requests.exceptions.RequestException as e:
        print(f"请求失败: {e}")
    except ValueError as e:
        print(f"JSON解析失败: {e}")
    except Exception as e:
        print(f"发生错误: {e}")

def get_video_id(bv):
    url = f'https://www.bilibili.com/video/{bv}'
    html = requests.get(url, headers=headers)
    html.encoding = 'utf-8'
    content = html.text
    aid_regx = '"aid":(.*?),"bvid":"{}"'.format(bv)
    video_aid = re.findall(aid_regx, content)[0]
    return video_aid


def fetch_comment_replies(video_id, comment_id, parent_user_name, max_pages=20):
    replies = []
    preLen = 0
    for page in range(1, max_pages + 1):
        url = f'https://api.bilibili.com/x/v2/reply/reply?oid={video_id}&type=1&root={comment_id}&ps=10&pn={page}'
        try:
            # 添加超时设置
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data and data.get('data') and data['data'].get('replies'):
                    for reply in data['data']['replies']:
                        reply_info = {
                            '用户昵称': reply['member']['uname'],
                            '评论内容': reply['content']['message'],
                            '被回复用户': parent_user_name,
                            '评论层级': '二级评论',
                            '性别': reply['member']['sex'],
                            '用户当前等级': reply['member']['level_info']['current_level'],
                            '点赞数量': reply['like'],
                            '回复时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(reply['ctime']))
                        }
                        replies.append(reply_info)
                    if preLen == len(replies):
                        break
                    preLen = len(replies)
                else:
                    return replies
        except requests.RequestException as e:
            print(f"请求出错: {e}")
            break
        # 控制请求频率
        time.sleep(1)
    return replies


def fetch_comments(video_id, max_pages=20):
    comments = []
    last_count = 0
    for page in range(1, max_pages + 1):
        url = f'https://api.bilibili.com/x/v2/reply?pn={page}&type=1&oid={video_id}&sort=2'
        try:
            # 添加超时设置
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                data = response.json()
                if data and data.get('data') and data['data'].get('replies'):
                    for comment in data['data']['replies']:
                        comment_info = {
                            '用户昵称': comment['member']['uname'],
                            '评论内容': comment['content']['message'],
                            '被回复用户': '',
                            '评论层级': '一级评论',
                            '性别': comment['member']['sex'],
                            '用户当前等级': comment['member']['level_info']['current_level'],
                            '点赞数量': comment['like'],
                            '回复时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(comment['ctime']))
                        }
                        comments.append(comment_info)
                        replies = fetch_comment_replies(video_id, comment['rpid'], comment['member']['uname'])
                        comments.extend(replies)
                if last_count == len(comments):
                    break
                last_count = len(comments)
            else:
                break
        except requests.RequestException as e:
            print(f"请求出错: {e}")
            break
        # 控制请求频率
        time.sleep(1)
    return comments

def get_unique_filename(base_name, extension=".xlsx"):
    base_name = base_name.split('.')[1]
    counter = 1
    filename = f".{base_name}{extension}"
    while os.path.exists(filename):
        filename = f".{base_name}_{counter}{extension}"
        counter += 1
    return filename

def save_comments_to_excel(comments, video_name, workbook, video_bv):
    if (video_bv in workbook.sheetnames):
        sheet = workbook[video_bv]
    else:
        sheet = workbook.create_sheet(title=video_bv)
        sheet.append(['用户昵称', '性别', '评论内容', '被回复用户', '评论层级', '用户当前等级', '点赞数量', '回复时间', video_name])
    
    for comment in comments:
        sheet.append([
            comment['用户昵称'], comment['性别'], comment['评论内容'], comment['被回复用户'],
            comment['评论层级'], comment['用户当前等级'], comment['点赞数量'], comment['回复时间']
        ])

def fetch_and_save_comments(video_name, video_bv, workbook):
    print(f'视频名字: {video_name}, video_bv: {video_bv}')
    aid = get_video_id(video_bv)
    video_id = aid
    comments = fetch_comments(video_id)
    save_comments_to_excel(comments, video_name, workbook, video_bv)

filename = './video_list.csv'
output_filename = './result/comment_output.xlsx'
max_threads = 5  # 可以调整的线程数

if __name__ == '__main__':

    #检测文件存在
    output_filename=get_unique_filename(output_filename)

    # 创建或加载工作簿
    if os.path.exists(output_filename):
        workbook = load_workbook(output_filename)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # 删除默认的Sheet

    #获取热点视频
    video_list = get_popular_videos()

    # 使用线程池并行抓取评论
    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        futures = [executor.submit(fetch_and_save_comments, video_name, video_bv, workbook) for video_name, video_bv in video_list]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"线程执行出错: {e}")

    # 保存工作簿
    workbook.save(output_filename)
